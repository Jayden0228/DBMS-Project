from cgitb import text
from distutils.cmd import Command
from tkinter import *
from tkinter import ttk
from typing import ItemsView
from PIL import ImageTk, Image
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter 
from tkinter import filedialog
import tkinter as tk
global COUNT
COUNT=1


def UP():
    global COUNT 
    textarea.delete(1.0,END)
   
    wb=load_workbook('C:/Users/Lloyd/Desktop/CLLoyd/Product Details/CraftozaProducts.xlsx')
    ws1=wb['Sheet1'];
    ws2=wb['Sheet2'];
    Uniq=ws2['J1'].value
 
    cat1={"Bamboo":"A","Coconut":"B","Clay":"C","Shells":"D","Wood":"E"}
    cat2={"Bag":"1","Home Deco":"2","Earthen pots":"3","Juwellry":"4","Funiture":"5"}

    if str(clicked1.get()) in cat1.keys():
        CAT1_Alpha=cat1[str(clicked1.get())]
        
    if str(clicked2.get()) in cat2.keys():
        CAT2_Alpha=cat2[str(clicked2.get())]

    UN="CRAF-"+str(Uniq)+"-"+CAT1_Alpha+"#"+CAT2_Alpha
   

    try:
        ws1.append([UN,str(pe1.get()),clicked1.get(),clicked2.get(),str(pe2.get()),str(pe3.get()),'=HYPERLINK("{}","{}")'.format(str(product_deets.filename),UN)])
        wb.save('C:/Users/Lloyd/Desktop/CLLoyd/Product Details/CraftozaProducts.xlsx')
        itemSection.insert("",'end',iid=COUNT,values=(UN,str(pe1.get()),clicked1.get(),clicked2.get(),str(pe2.get()),str(pe3.get())))
        ws2['J1']=Uniq+1
        COUNT=COUNT+1
        textarea.insert(END," Sheet Updated!")
        pe1.delete(0,END)
        pe2.delete(0,END)
        pe3.delete(0,END)

    except PermissionError:
        textarea.insert(END," Close the excel sheet")

 
            
def imageFunc():
    product_deets.filename=filedialog.askopenfilename(initialdir="C:/Users/Lloyd/Desktop/CLLoyd/Cropped Prod Images",title="Select a File",filetypes=(("png files","*.png"),("all files","*.*")))
    textarea.delete(1.0,END)
    textarea.insert(END,"Image selected")








root=Tk()
root.title('CRAFTOZA Excel Entry')
root.geometry('1366x768')
root.configure(bg='#e5e5e5')
colour='#34495E'

Main_Title=Frame(root,height=200,bg="white")
Main_Title.pack(fill=X)

imgg=Image.open("C:/Users/Lloyd/Desktop/CLLoyd/Images/Logo.png")
resizeIMG=imgg.resize((170,45))
img=ImageTk.PhotoImage(resizeIMG)
label=Label(Main_Title,image=img,bg="white")
label.pack()
Main_Below=Label(root,bg=colour)
Main_Below.pack(fill=X)
Main_Below=Label(root,bg='#F1C40F',height=1)
Main_Below.pack(fill=X)






product_deets=LabelFrame(root,text='Product Details',font=('Trebuchet MS',12),bg=colour,fg='white',pady=10)
product_deets.place(x=20,y=120,width=630,height=600)
textarea=Text(product_deets,width=25,height=3)
textarea.place(x=280,y=400)
textarea.insert(END," STATUS BOX:")

p1=Label(product_deets,text='              Product Name:  ',bg=colour,fg='white' ,font=('Microsoft JhengHei',12,'bold'),justify=LEFT,pady=10,)
p1.grid(row=0,column=0)
pe1=Entry(product_deets,width=50,relief=SUNKEN)
pe1.grid(row=0,column=1)

p2=Label(product_deets,text='             Price :  ',bg=colour,fg='white' ,font=('Microsoft JhengHei',12,'bold'),justify=LEFT,pady=10)
p2.grid(row=1,column=0)
pe2=Entry(product_deets,width=50,relief=SUNKEN,)
pe2.grid(row=1,column=1)

p3=Label(product_deets,text='             Seller :  ',bg=colour,fg='white' ,font=('Microsoft JhengHei',12,'bold'),justify=LEFT,pady=10)
p3.grid(row=2,column=0)
pe3=Entry(product_deets,width=50,relief=SUNKEN,)
pe3.grid(row=2,column=1)


p4=Label(product_deets,text='              Category (Type-1):  ',bg=colour,fg='white' ,font=('Microsoft JhengHei',12,'bold'),justify=LEFT,pady=10,)
p4.place(x=10,y=150)
clicked1=StringVar()
clicked1.set("Bamboo")
drop1=OptionMenu(product_deets,clicked1,"Bamboo","Clay","Coconut","Shells","Wood")
drop1.place(x=280,y=155)

p5=Label(product_deets,text='              Category (Type-2):  ',bg=colour,fg='white' ,font=('Microsoft JhengHei',12,'bold'),justify=LEFT,pady=10,)
p5.place(x=10,y=200)
clicked2=StringVar()
clicked2.set("Bag")
drop2=OptionMenu(product_deets,clicked2,"Bag","Home Deco","Earthen pots","Juwellry","Funiture")
drop2.place(x=280,y=210)


p6=Label(product_deets,text='              Add Image:  ',bg=colour,fg='white' ,font=('Microsoft JhengHei',12,'bold'),justify=LEFT,pady=10,)
p6.place(x=10,y=280)
Add_item1=Button(product_deets,text='Add Image',padx=20,pady=10,font=('Microsoft JhengHei',8,'bold'),command=imageFunc)
Add_item1.place(x=280,y=280)


Add_item2=Button(product_deets,text='Add Item ',padx=20,pady=10,font=('Microsoft JhengHei',8,'bold'),command=UP)
Add_item2.place(x=80,y=400)




style=ttk.Style()
style.configure("mystyle.Treeview.Heading",font=('Microsoft JhengHei',10,'bold'))
style.configure("mystyle.Treeview",font=('Microsoft JhengHei',9))
itemSection=ttk.Treeview(root,style="mystyle.Treeview")
itemSection['columns']=("ID","PRODUCT NAME","TYPE1","TYPE2","PRICE","SELLER")
itemSection.column("#0",anchor=W,width=0,stretch=NO)
itemSection.column("ID",anchor=W,width=40)
itemSection.column("PRODUCT NAME",anchor=W,width=200)
itemSection.column("TYPE1",anchor=W,width=100)
itemSection.column("TYPE2",anchor=W,width=80)
itemSection.column("PRICE",anchor=W,width=80)
itemSection.column("SELLER",anchor=W,width=80)


itemSection.heading("ID",text="ID",anchor=W)
itemSection.heading("PRODUCT NAME",text="PRODUCT NAME",anchor=W)
itemSection.heading("TYPE1",text="TYPE1",anchor=W)
itemSection.heading("TYPE2",text="TYPE2",anchor=W)
itemSection.heading("PRICE",text="PRICE",anchor=W)
itemSection.heading("SELLER",text="SELLER",anchor=W)

itemSection.place(x=680,y=120,width=630,height=600)

root.mainloop()