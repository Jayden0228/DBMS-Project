from cgitb import text
from distutils.cmd import Command
from tkinter import *
from tkinter import ttk
from typing import ItemsView
from PIL import ImageTk, Image
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter 
from tkinter import filedialog
import tkinter as tk
global COUNT
COUNT=1

global i
i=0
class CraftozaExcelUpdater:

     def UP(self):
        global COUNT 
        self.textarea.delete(1.0,END)
    
        self.wb=load_workbook('C:/Users/Lloyd/Desktop/CLLoyd/Product Details/CraftozaProducts.xlsx')
        self.ws1=self.wb['Sheet1'];
        self.ws2=self.wb['Sheet2'];
        self.Uniq=self.ws2['J1'].value
    
        self.cat1={"Bamboo":"A","Coconut":"B","Clay":"C","Shells":"D","Wood":"E"}
        self.cat2={"Bag":"1","Home Deco":"2","Earthen pots":"3","Juwellry":"4","Funiture":"5"}

        if str(self.clicked1.get()) in self.cat1.keys():
            self.CAT1_Alpha=self.cat1[str(self.clicked1.get())]
            
        if str(self.clicked2.get()) in self.cat2.keys():
            self.CAT2_Alpha=self.cat2[str(self.clicked2.get())]

        self.UN="CRAF-"+str(self.Uniq)+"-"+self.CAT1_Alpha+"#"+self.CAT2_Alpha
    

        try:
            self.ws1.append([self.UN,str(self.pe1.get()),self.clicked1.get(),self.clicked2.get(),str(self.pe2.get()),str(self.pe3.get()),'=HYPERLINK("{}","{}")'.format(str(self.product_deets.filename),self.UN)])
            self.ws2['J1']=self.Uniq+1
            self.wb.save('C:/Users/Lloyd/Desktop/CLLoyd/Product Details/CraftozaProducts.xlsx')
            self.itemSection.insert("",'end',iid=COUNT,values=(self.UN,str(self.pe1.get()),self.clicked1.get(),self.clicked2.get(),str(self.pe2.get()),str(self.pe3.get())))
            COUNT=COUNT+1
            self.textarea.insert(END," Sheet Updated!")
            self.pe1.delete(0,END)
            self.pe2.delete(0,END)
            self.pe3.delete(0,END)

        except PermissionError:
            self.textarea.insert(END," Close the excel sheet")

    
            
     def imageFunc(self):
        self.product_deets.filename=filedialog.askopenfilename(initialdir="C:/Users/Lloyd/Desktop/CLLoyd/Cropped Prod Images",title="Select a File",filetypes=(("png files","*.png"),("all files","*.*")))
        self.textarea.delete(1.0,END)
        self.textarea.insert(END,"Image selected")
   
   
     def __init__(self,master):

         self.Main_Title=Frame(master,height=200,bg="white")
         self.Main_Title.pack(fill=X)
         self.colour='#34495E'
         self.imgg=Image.open("C:/Users/Lloyd/Desktop/CLLoyd/Images/Logo.png")
         self.resizeIMG= self.imgg.resize((170,45))
         self.img=ImageTk.PhotoImage( self.resizeIMG)
         self.label=Label(self.Main_Title,image= self.img,bg="white")
         self.label.pack()
         self.Main_Below=Label(master,bg=self.colour)
         self.Main_Below.pack(fill=X)
         self.Main_Below=Label(master,bg='#F1C40F',height=1)
         self.Main_Below.pack(fill=X)






         self.product_deets=LabelFrame( master,text='Product Details',font=('Trebuchet MS',12),bg=self.colour,fg='white',pady=10)
         self.product_deets.place(x=20,y=120,width=630,height=600)
         self.textarea=Text( self.product_deets,width=25,height=3)
         self.textarea.place(x=280,y=400)
         self.textarea.insert(END," STATUS BOX:")

         self.p1=Label( self.product_deets,text='              Product Name:  ',bg=self.colour,fg='white' ,font=('Microsoft JhengHei',12,'bold'),justify=LEFT,pady=10,)
         self.p1.grid(row=0,column=0)
         self.pe1=Entry( self.product_deets,width=50,relief=SUNKEN)
         self.pe1.grid(row=0,column=1)

         self.p2=Label( self.product_deets,text='             Price :  ',bg=self.colour,fg='white' ,font=('Microsoft JhengHei',12,'bold'),justify=LEFT,pady=10)
         self.p2.grid(row=1,column=0)
         self.pe2=Entry(self.product_deets,width=50,relief=SUNKEN,)
         self.pe2.grid(row=1,column=1)

         self.p3=Label(self.product_deets,text='             Seller :  ',bg=self.colour,fg='white' ,font=('Microsoft JhengHei',12,'bold'),justify=LEFT,pady=10)
         self.p3.grid(row=2,column=0)
         self.pe3=Entry(self.product_deets,width=50,relief=SUNKEN,)
         self.pe3.grid(row=2,column=1)


         self.p4=Label(self.product_deets,text='              Category (Type-1):  ',bg=self.colour,fg='white' ,font=('Microsoft JhengHei',12,'bold'),justify=LEFT,pady=10,)
         self.p4.place(x=10,y=150)
         self.clicked1=StringVar()
         self.clicked1.set("Bamboo")
         self.drop1=OptionMenu(self.product_deets,self.clicked1,"Bamboo","Clay","Coconut","Shells","Wood")
         self.drop1.place(x=280,y=155)
 
         self.p5=Label(self.product_deets,text='              Category (Type-2):  ',bg=self.colour,fg='white' ,font=('Microsoft JhengHei',12,'bold'),justify=LEFT,pady=10,)
         self.p5.place(x=10,y=200)
         self.clicked2=StringVar()
         self.clicked2.set("Bag")
         self.drop2=OptionMenu(self.product_deets,self.clicked2,"Bag","Home Deco","Earthen pots","Juwellry","Funiture")
         self.drop2.place(x=280,y=210)


         self.p6=Label(self.product_deets,text='              Add Image:  ',bg=self.colour,fg='white' ,font=('Microsoft JhengHei',12,'bold'),justify=LEFT,pady=10,)
         self.p6.place(x=10,y=280)
         self.Add_item1=Button(self.product_deets,text='Add Image',padx=20,pady=10,font=('Microsoft JhengHei',8,'bold'),command=self.imageFunc)
         self.Add_item1.place(x=280,y=280)


         self.Add_item2=Button(self.product_deets,text='Add Item ',padx=20,pady=10,font=('Microsoft JhengHei',8,'bold'),command=self.UP)
         self.Add_item2.place(x=80,y=400)




         self.style=ttk.Style()
         self. style.configure("mystyle.Treeview.Heading",font=('Microsoft JhengHei',10,'bold'))
         self.style.configure("mystyle.Treeview",font=('Microsoft JhengHei',9))
         self.itemSection=ttk.Treeview(master,style="mystyle.Treeview")
         self.itemSection['columns']=("ID","PRODUCT NAME","TYPE1","TYPE2","PRICE","SELLER")
         self.itemSection.column("#0",anchor=W,width=0,stretch=NO)
         self.itemSection.column("ID",anchor=W,width=40)
         self.itemSection.column("PRODUCT NAME",anchor=W,width=200)
         self.itemSection.column("TYPE1",anchor=W,width=100)
         self.itemSection.column("TYPE2",anchor=W,width=80)
         self.itemSection.column("PRICE",anchor=W,width=80)
         self.itemSection.column("SELLER",anchor=W,width=80)


         self.itemSection.heading("ID",text="ID",anchor=W)
         self.itemSection.heading("PRODUCT NAME",text="PRODUCT NAME",anchor=W)
         self.itemSection.heading("TYPE1",text="TYPE1",anchor=W)
         self.itemSection.heading("TYPE2",text="TYPE2",anchor=W)
         self.itemSection.heading("PRICE",text="PRICE",anchor=W)
         self.itemSection.heading("SELLER",text="SELLER",anchor=W)

         self.itemSection.place(x=680,y=120,width=630,height=600)


      
         self.EntryFields=[self.pe1,self.pe2,self.pe3]
         
         master.bind('<Down>',self.downKey)
         master.bind('<Up>',self.upKey)



     def downKey(self,event):
                global i
                i=i+1
                self.EntryFields[i].focus()

     def upKey(self,event):
                global i
                i=i-1
                self.EntryFields[i].focus()
        
                

        



if __name__ == '__main__':
    root=Tk()
    root.title('CRAFTOZA Excel Entry')
    root.geometry('1366x768')
    root.configure(bg='#e5e5e5')
      
    obj=CraftozaExcelUpdater(root)
    root.mainloop()
    
